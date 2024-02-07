import openpyxl
import shutil
import os

import warnings
warnings.simplefilter("ignore")

def 合并Excel文件(输入文件夹, 输出文件):
    # 创建一个新的工作簿
    输出工作簿 = openpyxl.Workbook()
    输出工作表 = 输出工作簿.active

#    for file_name in os.listdir(输入文件夹):
#        print("file_name: " + file_name)
        
        
    # 遍历输入文件夹中的所有文件
    for 文件名 in os.listdir(输入文件夹):
        if (文件名.endswith(".xlsx") or 文件名.endswith(".xls")) and (文件名 != "merged_data.xlsx"):
            文件路径 = os.path.join(输入文件夹, 文件名)

            # 打开每个输入文件
            输入工作簿 = openpyxl.load_workbook(文件路径)
            输入工作表 = 输入工作簿.active

            # 遍历输入工作表中的每一行，并将其复制到输出工作表中
            for 行 in 输入工作表.iter_rows(min_row=1, max_row=输入工作表.max_row, values_only=True):
                输出工作表.append(行)

            # 关闭输入工作簿
            输入工作簿.close()
        else:print(文件名)
    # 保存合并后的工作簿到输出文件
    输出工作簿.save(输出文件)
    print(f'合并后的Excel文件已保存到 {输出文件}')

# 指定包含Excel文件的输入文件夹和输出文件路径
输入文件夹 = "/Users/boge/hzw/git/My_Study/My_Study_New/Python/MergedExcel"
输出文件 = "/Users/boge/hzw/git/My_Study/My_Study_New/Python/MergedExcel/merged_data.xlsx"

# 调用函数来合并Excel文件
合并Excel文件(输入文件夹, 输出文件)
