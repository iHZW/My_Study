import openpyxl
import os
import warnings
warnings.simplefilter("ignore")

# 创建一个新的工作簿
merged_workbook = openpyxl.Workbook()
merged_sheet = merged_workbook.active

# 设置包含要合并的所有Excel文件的文件夹路径
folder_path = '/Users/boge/hzw/git/My_Study/My_Study_New/Python/MergedExcel'

# 获取文件夹中的所有Excel文件
all_files = [file for file in os.listdir(folder_path) if (file.endswith('.xlsx') or file.endswith('.xsl') and (file != 'merged_data.xlsx'))]

# 循环处理每个Excel文件
for file in all_files:
    file_path = os.path.join(folder_path, file)
#    print("file_path: " + file_path)
    # 打开每个Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择要复制的工作表（这里选择第一个工作表，你可以根据需要更改）
    sheet_to_copy = workbook.active

    # 复制数据到合并的工作表中
    for row in sheet_to_copy.iter_rows(values_only=True):
        merged_sheet.append(row)

# 保存合并后的工作簿为新的Excel文件
output_file_path = '/Users/boge/hzw/git/My_Study/My_Study_New/Python/MergedExcel/merged_data.xlsx'
merged_workbook.save(output_file_path)

print("合并完成，结果保存在:", output_file_path)
